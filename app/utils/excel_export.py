from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO


def generate_tmt_excel(data, sheet_name="TMT DPR REPORT"):

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    headers = [
        "Sr",
        "Site Code",
        "Site Name",
        "PO Qty",
        "Supplied",
        "Balance",
        "Supply %",
        "Status",
        "PM Comment",
        "Director Remarks"
    ]

    # Header Style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1661BE", end_color="1661BE", fill_type="solid")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_num, item in enumerate(data, 2):

        ws.cell(row=row_num, column=1).value = item.get("serial") or 0
        ws.cell(row=row_num, column=2).value = item.get("site_code") or "-"
        ws.cell(row=row_num, column=3).value = item.get("site_name") or "-"
        ws.cell(row=row_num, column=4).value = item.get("po_qty") or 0
        ws.cell(row=row_num, column=5).value = item.get("supplied") or 0
        ws.cell(row=row_num, column=6).value = item.get("balance") or 0
        ws.cell(row=row_num, column=7).value = f"{item.get('supply_percent') or 0}%"
        ws.cell(row=row_num, column=8).value = item.get("pm_remarks") or "-"
        ws.cell(row=row_num, column=9).value = item.get("pm_comment") or "-"    
        ws.cell(row=row_num, column=10).value = item.get("director_remarks") or "-"

    # Auto column width (simple)
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 5

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output